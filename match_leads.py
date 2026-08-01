#!/usr/bin/env python3
"""
Facebook leads  <->  HubSpot contacts reconciliation.

Reads a Facebook lead export (Excel), fetches every HubSpot contact created in a
chosen date range, and reports which Facebook leads are NOT in HubSpot (missing
because of a broken sync). Matching is by email and phone, both normalised so
different formats still match.

Runs on GitHub Actions. Inputs come from environment variables (set by the workflow):
  HUBSPOT_TOKEN  - HubSpot private-app token (from the repo secret)
  START_DATE     - YYYY-MM-DD  (leave blank = first day of this month)
  END_DATE       - YYYY-MM-DD  (leave blank = today)
  LEADS_FILE     - Excel filename in the repo (default: facebook-leads.xlsx)
  OUT_FILE       - output filename (default: unmatched-leads.xlsx)
Optional column overrides (only if auto-detection picks the wrong column):
  COL_EMAIL, COL_PHONE, COL_NAME, COL_DATE  - exact header text from your Excel
"""

import os
import re
import io
import csv
import sys
import glob
import time
import datetime
import requests
from openpyxl import load_workbook, Workbook

# ------------------------------------------------------------------ config
TOKEN      = os.environ.get("HUBSPOT_TOKEN", "").strip()
LEADS_FILE = os.environ.get("LEADS_FILE", "facebook-leads.xlsx").strip()
OUT_FILE   = os.environ.get("OUT_FILE", "unmatched-leads.xlsx").strip()
START_DATE = os.environ.get("START_DATE", "").strip()
END_DATE   = os.environ.get("END_DATE", "").strip()

OV_EMAIL = os.environ.get("COL_EMAIL", "").strip()
OV_PHONE = os.environ.get("COL_PHONE", "").strip()
OV_NAME  = os.environ.get("COL_NAME", "").strip()
OV_DATE  = os.environ.get("COL_DATE", "").strip()

BASE = "https://api.hubapi.com"
HDRS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
SEARCH_THROTTLE = 0.3   # seconds to wait after each HubSpot search (stay under the per-second limit)

# HubSpot properties that may hold a phone number (add custom ones here if needed)
HS_PHONE_PROPS = ["phone", "mobilephone"]

# ------------------------------------------------------------------ helpers
def log(msg):
    print(msg, flush=True)

def summary(md):
    path = os.environ.get("GITHUB_STEP_SUMMARY")
    if path:
        with open(path, "a", encoding="utf-8") as f:
            f.write(md + "\n")

def die(msg):
    log("ERROR: " + msg)
    summary(f"### ❌ Lead match failed\n\n{msg}\n")
    sys.exit(1)

def norm_email(v):
    if not v:
        return ""
    return str(v).strip().lower()

def norm_phone(v):
    digits = re.sub(r"\D", "", str(v or ""))
    return digits[-9:] if len(digits) >= 9 else digits   # last 9 digits = the mobile's unique part

def month_defaults():
    today = datetime.date.today()
    first = today.replace(day=1)
    return first.isoformat(), today.isoformat()

def parse_date(s):
    """Best-effort date-only parse; returns a date or None."""
    if s is None or s == "":
        return None
    if isinstance(s, datetime.datetime):
        return s.date()
    if isinstance(s, datetime.date):
        return s
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(s))
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # try common d/m/Y or m/d/Y
    m = re.search(r"(\d{1,2})[/](\d{1,2})[/](\d{4})", str(s))
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(y, b, a) if a > 12 else datetime.date(y, a, b)
        except ValueError:
            return None
    return None

def to_ms(date_str, end=False):
    d = datetime.datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
    if end:
        d = d + datetime.timedelta(days=1) - datetime.timedelta(milliseconds=1)
    return int(d.timestamp() * 1000)

# ------------------------------------------------------------------ Excel
def find_col(headers, override, keywords):
    lowered = [(h or "").strip().lower() for h in headers]
    if override:
        for i, h in enumerate(headers):
            if (h or "").strip().lower() == override.strip().lower():
                return i
        die(f'Column "{override}" was not found in the file. Headers are: {headers}')
    for kw in keywords:  # keywords are tried in priority order
        for i, h in enumerate(lowered):
            if kw in h:
                return i
    return None

def resolve_leads_file():
    """Use LEADS_FILE if it exists; otherwise find a Facebook export in the repo."""
    if LEADS_FILE and os.path.exists(LEADS_FILE):
        return LEADS_FILE
    for cand in ["facebook-leads.csv", "facebook-leads.xlsx",
                 "facebook_leads.csv", "facebook_leads.xlsx", "leads.csv", "leads.xlsx"]:
        if os.path.exists(cand):
            return cand
    found = sorted(glob.glob("*.csv") + glob.glob("*.xlsx"))
    found = [f for f in found if os.path.basename(f) != OUT_FILE]
    if len(found) == 1:
        return found[0]
    die(f'Could not find your leads file. Upload it to the repo root as "facebook-leads.csv" '
        f'or "facebook-leads.xlsx". Files seen: {found or "none"}.')

def load_table(path):
    """Return (headers, rows) for a .csv or .xlsx file. Cells are strings/numbers/dates."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        raw = open(path, "rb").read()
        if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):        # UTF-16 (Facebook's usual export)
            text = raw.decode("utf-16", errors="replace")
        elif raw[:3] == b"\xef\xbb\xbf":                 # UTF-8 with BOM
            text = raw.decode("utf-8-sig", errors="replace")
        else:
            try:
                text = raw.decode("utf-8")
            except UnicodeDecodeError:
                text = raw.decode("latin-1", errors="replace")
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        first = next((ln for ln in text.split("\n") if ln.strip()), "")
        counts = {d: first.count(d) for d in ["\t", ",", ";", "|"]}
        delim = max(counts, key=counts.get)
        if counts[delim] == 0:
            delim = ","
        all_rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        if not all_rows:
            die("The file appears to be empty.")
        return [str(h).strip() for h in all_rows[0]], all_rows[1:]
    elif ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            die("The Excel file appears to be empty.")
        rows = [list(r) for r in it]
        wb.close()
        return headers, rows
    else:
        die(f'Unsupported file type "{ext}". Please upload a .csv or .xlsx file.')

def read_leads():
    path = resolve_leads_file()
    log(f"Reading leads from: {path}")
    headers, data_rows = load_table(path)

    i_email = find_col(headers, OV_EMAIL, ["email", "e-mail", "mail"])
    i_phone = find_col(headers, OV_PHONE, ["phone", "mobile", "whatsapp", "contact number", "number"])
    i_name  = find_col(headers, OV_NAME,  ["full name", "full_name", "name"])
    i_first = find_col(headers, "", ["first name", "first_name"])
    i_last  = find_col(headers, "", ["last name", "last_name"])
    i_date  = find_col(headers, OV_DATE,  ["created_time", "created time", "created", "submitted", "date", "time"])

    if i_email is None and i_phone is None:
        die(f"Could not find an email or phone column. Headers are: {headers}. "
            f"Set COL_EMAIL / COL_PHONE to the exact header text if needed.")

    log(f"Detected columns -> email: {headers[i_email] if i_email is not None else '(none)'} | "
        f"phone: {headers[i_phone] if i_phone is not None else '(none)'} | "
        f"name: {headers[i_name] if i_name is not None else '(first+last or none)'} | "
        f"date: {headers[i_date] if i_date is not None else '(none)'}")

    leads = []
    for r in data_rows:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        def cell(idx):
            return r[idx] if idx is not None and idx < len(r) else None
        email = norm_email(cell(i_email))
        phone = norm_phone(cell(i_phone))
        if i_name is not None:
            name = str(cell(i_name) or "").strip()
        else:
            name = (str(cell(i_first) or "").strip() + " " + str(cell(i_last) or "").strip()).strip()
        d = parse_date(cell(i_date)) if i_date is not None else None
        leads.append({
            "name": name,
            "email_raw": str(cell(i_email) or "").strip(),
            "phone_raw": str(cell(i_phone) or "").strip(),
            "email": email,
            "phone": phone,
            "date": d,
        })
    return leads, (i_date is not None)

# ------------------------------------------------------------------ HubSpot
def hs_search(body):
    url = f"{BASE}/crm/v3/objects/contacts/search"
    for attempt in range(8):
        r = requests.post(url, headers=HDRS, json=body, timeout=60)
        if r.status_code == 429:
            wait = 0
            ra = r.headers.get("Retry-After")
            if ra:
                try:
                    wait = float(ra)
                except ValueError:
                    wait = 0
            if wait <= 0:
                wait = 2 + attempt * 2
            wait = min(wait, 15)
            log(f"HubSpot rate limit hit; waiting {wait:.0f}s then retrying (attempt {attempt + 1}/8)...")
            time.sleep(wait)
            continue
        if r.status_code == 401:
            die("HubSpot rejected the token (401). Check the HUBSPOT_TOKEN secret and its scopes.")
        if r.status_code >= 300:
            die(f"HubSpot API error {r.status_code}: {r.text[:300]}")
        time.sleep(SEARCH_THROTTLE)
        return r.json()
    die("HubSpot kept rate-limiting the requests. This usually means the token is busy with "
        "your other scripts, or the workflow was run several times in quick succession. "
        "Wait a minute and run it again.")

def fetch_contacts_in_range(start_ms, end_ms):
    """All contacts created in the window -> normalised email set + phone set."""
    emails, phones, total = set(), set(), 0
    after = None
    props = ["email", "createdate"] + HS_PHONE_PROPS
    while True:
        body = {
            "filterGroups": [{"filters": [
                {"propertyName": "createdate", "operator": "BETWEEN", "value": start_ms, "highValue": end_ms}
            ]}],
            "properties": props,
            "sorts": [{"propertyName": "createdate", "direction": "ASCENDING"}],
            "limit": 100,
        }
        if after:
            body["after"] = after
        data = hs_search(body)
        for c in data.get("results", []):
            total += 1
            p = c.get("properties", {})
            e = norm_email(p.get("email"))
            if e:
                emails.add(e)
            for pp in HS_PHONE_PROPS:
                ph = norm_phone(p.get(pp))
                if ph:
                    phones.add(ph)
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break
        if total >= 9900:
            log("WARNING: reached the ~10,000-contact limit for this range. "
                "Use a shorter date range to be fully accurate.")
            break
    return emails, phones, total

def emails_exist_anywhere(email_list):
    """Global check (ignores date): which of these emails exist in HubSpot at all."""
    found = set()
    email_list = [e for e in email_list if e]
    for i in range(0, len(email_list), 100):
        batch = email_list[i:i + 100]
        after = None
        while True:
            body = {
                "filterGroups": [{"filters": [
                    {"propertyName": "email", "operator": "IN", "values": batch}
                ]}],
                "properties": ["email"],
                "limit": 100,
            }
            if after:
                body["after"] = after
            data = hs_search(body)
            for c in data.get("results", []):
                e = norm_email(c.get("properties", {}).get("email"))
                if e:
                    found.add(e)
            after = data.get("paging", {}).get("next", {}).get("after")
            if not after:
                break
    return found

# ------------------------------------------------------------------ main
def main():
    if not TOKEN:
        die("HUBSPOT_TOKEN is not set. Add it as a repository secret.")

    start = START_DATE or month_defaults()[0]
    end   = END_DATE or month_defaults()[1]
    try:
        start_ms, end_ms = to_ms(start), to_ms(end, end=True)
    except ValueError:
        die("Dates must be in YYYY-MM-DD format, e.g. 2026-07-01.")
    if start_ms > end_ms:
        die("Start date is after end date.")

    log(f"Range: {start} to {end}")
    leads, has_date = read_leads()
    log(f"Read {len(leads)} rows.")

    # keep only leads whose Facebook date falls in the range (if a date column exists)
    if has_date:
        in_range, out_range = [], 0
        s_d = datetime.date.fromisoformat(start)
        e_d = datetime.date.fromisoformat(end)
        for ld in leads:
            if ld["date"] is None or (s_d <= ld["date"] <= e_d):
                in_range.append(ld)
            else:
                out_range += 1
        leads = in_range
        log(f"{len(leads)} Facebook leads fall in the selected range ({out_range} outside were skipped).")

    log("Fetching HubSpot contacts created in the range...")
    hs_emails, hs_phones, hs_total = fetch_contacts_in_range(start_ms, end_ms)
    log(f"HubSpot has {hs_total} contacts created in this range.")

    # ---- categorise: created-in-range match / older match / truly missing ----
    matched_in, prelim_unmatched, skipped = [], [], []
    for ld in leads:
        if not ld["email"] and not ld["phone"]:
            skipped.append(ld)
            continue
        if (ld["email"] and ld["email"] in hs_emails) or (ld["phone"] and ld["phone"] in hs_phones):
            matched_in.append(ld)
        else:
            prelim_unmatched.append(ld)

    found_anywhere = set()
    recheck = [ld["email"] for ld in prelim_unmatched if ld["email"]]
    if recheck:
        log(f"{len(prelim_unmatched)} leads have no HubSpot contact created in this range. "
            f"Checking whether they exist in HubSpot from another time...")
        found_anywhere = emails_exist_anywhere(recheck)

    matched_outside, missing = [], []
    for ld in prelim_unmatched:
        if ld["email"] and ld["email"] in found_anywhere:
            matched_outside.append(ld)      # exists in HubSpot, just created outside the window
        else:
            missing.append(ld)              # not found in HubSpot at all

    checked = len(matched_in) + len(matched_outside) + len(missing)

    # ---- write the results workbook (one sheet per bucket) ----
    def add_sheet(wb, title, rows_, first=False):
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
        ws.append(["Name", "Email", "Phone", "Facebook date"])
        for ld in rows_:
            ws.append([ld["name"], ld["email_raw"], ld["phone_raw"],
                       ld["date"].isoformat() if ld["date"] else ""])

    wb = Workbook()
    add_sheet(wb, "Missing from HubSpot", missing, first=True)
    add_sheet(wb, "In HubSpot but older", matched_outside)
    if skipped:
        add_sheet(wb, "Skipped (no email or phone)", skipped)
    wb.save(OUT_FILE)

    # ---- console + step summary ----
    log("=" * 52)
    log(f"Facebook leads checked:            {checked}")
    log(f"In HubSpot (created this range):   {len(matched_in)}")
    log(f"In HubSpot (created another time): {len(matched_outside)}")
    log(f"MISSING from HubSpot entirely:     {len(missing)}")
    if skipped:
        log(f"Skipped (no email/phone):          {len(skipped)}")
    log("=" * 52)

    lines = [
        "## Facebook to HubSpot lead match",
        "",
        f"**Range:** {start} to {end}",
        "",
        "| | Count |",
        "|---|---:|",
        f"| Facebook leads checked | {checked} |",
        f"| In HubSpot, created in this range | {len(matched_in)} |",
        f"| In HubSpot, but created another time | {len(matched_outside)} |",
        f"| **Missing from HubSpot entirely** | **{len(missing)}** |",
    ]
    if skipped:
        lines.append(f"| Skipped (no email/phone in row) | {len(skipped)} |")
    lines += ["", f"For reference, HubSpot has {hs_total} contacts created in this range.", ""]

    def table(title, rows_):
        out = [f"### {title} (first 15 — full list in the downloaded file)", "",
               "| Name | Email | Phone | FB date |", "|---|---|---|---|"]
        for ld in rows_[:15]:
            out.append(f"| {ld['name'] or '-'} | {ld['email_raw'] or '-'} | "
                       f"{ld['phone_raw'] or '-'} | {ld['date'].isoformat() if ld['date'] else '-'} |")
        out.append("")
        return out

    if missing:
        lines += table("Missing from HubSpot", missing)
    else:
        lines += ["✅ Every Facebook lead in this range was found in HubSpot — nothing is truly missing.", ""]
    if matched_outside:
        lines += table("In HubSpot, but created outside this range", matched_outside)
        lines += ["These already exist in HubSpot from another time (returning people or older records). "
                  "Review them if you expected a brand-new contact this period.", ""]
    lines.append(f"⬇️ Download **{OUT_FILE}** from the Artifacts section for the full lists.")
    summary("\n".join(lines))

if __name__ == "__main__":
    main()
